import requests
from openpyxl import load_workbook
from xls2xlsx import XLS2XLSX
prefix_nome_arquivo_saida='func_caucaia_'

###Definindo Função para criar link dos aquivos desejados###
def url_mes_ano (mes, ano):
    mes=str(mes)
    ano=str(ano)
    api_url_prefix='https://api.portaltransparencia.sead.caucaia.ce.gov.br/funcionario/excel/1/'
    api_url_final=api_url_prefix + mes + '/' + ano
    return api_url_final

#criando o arquivo de saída
def download_arquivo (mes, ano):
    print(f'Enviando requisição de arquivo do mes:{mes} e do anos:{ano} ')
    resposta_da_requisicao=requests.get(url_mes_ano(mes,ano))
    if resposta_da_requisicao.status_code == 200:
        print('A requisição foi um sucesso!')
        with open(prefix_nome_arquivo_saida+mes+'_'+ano+'.xls','wb') as arquivo_de_saida:
            arquivo_de_saida.write(resposta_da_requisicao.content)
    else:
        print('Não foi possível realizar a requisição')

#CONVERTER PLANILHA PARA XLSX
def convert_para_xlsx(mes, ano):
    print('Convertendo arquivos xls em xlsx...')
    planilha01=XLS2XLSX(prefix_nome_arquivo_saida+mes+'_'+ano+'.xls')
    planilha01.to_xlsx(prefix_nome_arquivo_saida+mes+'_'+ano+'.xlsx')

mes1=str(input('Insira o Primeiro mes para a comparação(Ex.: 01,02,03...11,12):'))
ano1=str(input('Insira o ano (Ex.: 2022,2023):'))
mes2=str(input('Insira o Segundo mes para a comparação(Ex.: 01,02,03...11,12):'))
ano2=str(input('Insira o ano (Ex.: 2022,2023):'))

download_arquivo(mes1,ano1)
download_arquivo(mes2,ano2)
convert_para_xlsx(mes1,ano1)
convert_para_xlsx(mes2,ano2)

print(f'Abrindo planilha {prefix_nome_arquivo_saida+mes1+'_'+ano1+'.xlsx'} ......')
planilha01= load_workbook(filename=prefix_nome_arquivo_saida+mes1+'_'+ano1+'.xlsx')
print(f'Abrindo ABA da planilha {prefix_nome_arquivo_saida+mes1+'_'+ano1+'.xlsx'}.......')
planilha01_aba= planilha01[planilha01.sheetnames[0]]
print('Planilha 01 carregada com sucesso!')
#coleta o numero de linhas da planilha01
planilha01_num_linhas=planilha01_aba.max_row

#abrir a segunda Planilha
print(f'Abrindo planilha {prefix_nome_arquivo_saida+mes2+'_'+ano2+'xlsx'}.....')
planilha02=load_workbook(filename=prefix_nome_arquivo_saida+mes2+'_'+ano2+'.xlsx')

print(f'Abrindo abra da planilha {prefix_nome_arquivo_saida+mes2+'_'+ano2+'xlsx'}.....')
planilha02_aba=planilha02[planilha02.sheetnames[0]]
print('Planilha 02 carregada com sucesso!')
#coleta o numero de linhas da planilha02
planilha02_num_linhas=planilha02_aba.max_row
print('Iniciando verificação.....')
#percorre todas as linhas da coluna B da planilha 01, excluindo as duas primeiras
for p1 in range(planilha01_num_linhas):
    #NOME A SER PESQUISADO:
    planilha01_valor=str(planilha01_aba[str('A'+str(p1+3))].value)
    #criando variavel de controle pra identificar se a informacao fi encontrada
    controle=0
    for p2 in range(planilha02_num_linhas):
        planilha02_valor=str(planilha02_aba[str('A'+str(p2+3))].value)
        
        #comparando as duas strings
        if planilha01_valor == planilha02_valor:
            controle=1
        
    if controle == 0:
        #filtra pelo cargo de professor
        #if planilha01_aba[str('E'+str(p1+3))].value == 'PROFESSOR':
        print(f'-{planilha01_aba[str('B'+str(p1+3))].value} de mes:{mes1} e ano:{ano1} NÃO ENCONTRADO em mes:{mes2} e ano:{ano2}!!!')

print(f'######### Fim do Script #########')
    
