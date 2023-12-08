from openpyxl import load_workbook
import pandas as pd

def ondeEsta(ws:object,texto='NOMBRE'):
    """
    Esta funcao busca pela primeira ocorrencia de um determinado texto em um objeto do tipo WorkSheet do Openpyxl.
    Por default esse texto é a palavra 'NOMBRE'
    """
    #buscar pelo valor 'Nombre' percorrendo as linhas
    for linha in ws[ws.dimensions]:
        for celula in linha:
            if celula.value == texto:
                return (celula.row, celula.column)
            
def coletaDadoAbaixoCelula(posicao:tuple,ws:object):
    # Coleta os valores abaixo do nome "Nombre" encontrado e armazenda em lista_de_nomes
    lista_valores = []
    for i in range(posicao[0]+1,len(ws.row_dimensions)):
        if (celula := ws.cell(row=i,column=posicao[1]).value) != None:
            lista_valores.append(celula)
    return lista_valores

def subdivisao(posicao:tuple,lista:list):

    coluna_limite , dados_saida = posicao[1], []

    for i in range(1,coluna_limite):#percorrendo colunas
        for l in range(1,len(ws.row_dimensions)):
            ((valor := ws.cell(row=l, column=i).value))#pega o valor de cada celula
            if valor and valor not in lista:
                valor01 = valor
            elif valor in lista:
                dados_saida.append([ valor01.strip()+';'+valor ])
    return dados_saida

print('Lendo arquivo...')
#abrindo arquivo
file_path = 'equipes4u/VSE-EQUIPES_macro_rev01_2.xlsm'

wb = load_workbook(filename = file_path)
print('Coletando lista de abas...')
#pegando lista de abas, porem, so pega as que tem tamanho igual a 6, e que sao numeros assim sabemos que o nome esta correto
lista_de_abas = [x  for x in wb.sheetnames if (len(x)==6 and x.isnumeric())]

df = pd.DataFrame({'projeto':[], 'data':[], 'equipe':[], 'nome':[]})

print('Salvando dados no DataFrame...')

for aba in lista_de_abas:
        #converter aba no formato de data
        data = aba[:2]+'/'+aba[2:4]+'/20'+aba[4:]
        #abrindo aba dentro da do arquivo
        ws = wb[aba]
        #buscando posicao no nome Nombre e armasenando em posicao
        posicao = ondeEsta(ws)
        #pegando lista das informacoes abaixo no NOmbre
        nomes = coletaDadoAbaixoCelula(posicao,ws)
        equipe_nome = subdivisao(posicao,nomes)
        for item in equipe_nome:
            equipe = item[0][:item[0].index(';')].strip() #pega o primeiro nome retira espacos no final
            if equipe.find(' ') > 0: #verifica se tem palavra composta
                equipe = equipe[:(equipe.find(' ')+1)].strip() #gardando somente uma palavra
            nome = item[0][item[0].index(';')+1:].strip() #pegando o segundo elemento apos a ;
            df.loc[len(df)] = {'projeto': 'VSE', 'data':data, 'equipe': equipe, 'nome':nome} #adicionar linha ao df
            
wb.close()
df.to_csv('equipes4u/saida.csv')#salvando arquivos

#salvando arquivos resumidos
df.loc[:, ["equipe", "nome", "data"]].\
    groupby(['equipe','nome']).count().\
    sort_values(by='equipe').rename(columns={'data':'freq'}).\
    to_csv('equipes4u/resumo.csv')

print('Finalizado')